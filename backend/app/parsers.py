"""
Parsers for multiple input formats:
  - JSON  (generic vulnerability schema)
  - Excel (.xlsx / .xls)
  - PDF   (text extraction + heuristic finding detection)
"""

import io
import json
import re
from typing import Any, Dict, List, Optional


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

_SEVERITY_ALIASES: Dict[str, str] = {
    "critical": "critical", "crit": "critical", "4": "critical",
    "high": "high", "3": "high",
    "medium": "medium", "med": "medium", "moderate": "medium", "2": "medium",
    "low": "low", "1": "low",
    "info": "info", "informational": "info", "information": "info",
    "none": "info", "0": "info",
}


def _normalize_severity(raw: Any) -> str:
    """Map various severity labels to the canonical set."""
    if raw is None:
        return "info"
    return _SEVERITY_ALIASES.get(str(raw).strip().lower(), "info")


def _safe_float(val: Any) -> Optional[float]:
    """Try to cast a value to float; return None on failure."""
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _make_finding(
    tool: str,
    title: str,
    severity: str,
    cvss: Any = None,
    asset: str = "",
    endpoint: Optional[str] = None,
    cwe: Optional[str] = None,
    owasp: Optional[str] = None,
    evidence: Optional[Dict] = None,
    raw: Optional[Any] = None,
) -> Dict[str, Any]:
    """Build a canonical normalized finding dict."""
    return {
        "tool": tool,
        "title": str(title or "Untitled Finding"),
        "severity": _normalize_severity(severity),
        "cvss": _safe_float(cvss),
        "asset": str(asset or ""),
        "endpoint": endpoint or None,
        "evidence": evidence or {},
        "cwe": cwe or None,
        "owasp": owasp or None,
        "raw": raw or {},
    }


# ---------------------------------------------------------------------------
# 1. JSON parser
# ---------------------------------------------------------------------------

def parse_generic_vuln_json(payload: Dict[str, Any]) -> List[Dict[str, Any]]:
    """
    Expected schema:
    {
      "tool": "nessus|burp|generic",
      "findings": [
        {
          "title": "...",
          "severity": "critical|high|medium|low|info",
          "cvss": 9.1,
          "asset": "api.prod",
          "endpoint": "/path",
          "evidence": { ... },
          "cwe": "CWE-284",
          "owasp": "A01:2021"
        }
      ]
    }
    """
    tool = payload.get("tool", "generic")
    findings = payload.get("findings", [])
    out = []
    for f in findings:
        out.append(_make_finding(
            tool=tool,
            title=f.get("title"),
            severity=f.get("severity"),
            cvss=f.get("cvss"),
            asset=f.get("asset"),
            endpoint=f.get("endpoint"),
            evidence=f.get("evidence", {}),
            cwe=f.get("cwe"),
            owasp=f.get("owasp"),
            raw=f,
        ))
    return out


# ---------------------------------------------------------------------------
# 2. Excel parser (.xlsx / .xls)
# ---------------------------------------------------------------------------

# Flexible column name mapping – we try many common header variations.
_COL_MAP: Dict[str, List[str]] = {
    "title":    ["title", "name", "finding", "vulnerability", "vuln", "issue",
                 "finding name", "vulnerability name", "issue name", "summary",
                 "plugin name", "check name", "rule name"],
    "severity": ["severity", "risk", "risk level", "risk rating", "priority",
                 "sev", "criticality", "threat level", "impact level", "level"],
    "cvss":     ["cvss", "cvss score", "cvss v3", "cvssv3", "cvss3",
                 "cvss v2", "cvssv2", "base score", "score", "risk score"],
    "asset":    ["asset", "host", "hostname", "ip", "ip address", "target",
                 "affected host", "affected asset", "server", "system",
                 "machine", "device"],
    "endpoint": ["endpoint", "url", "path", "uri", "port", "service",
                 "affected url", "location", "resource"],
    "cwe":      ["cwe", "cwe id", "cwe-id", "weakness"],
    "owasp":    ["owasp", "owasp category", "owasp id", "owasp top 10"],
    "evidence": ["evidence", "details", "description", "proof",
                 "observation", "output", "plugin output", "synopsis",
                 "solution", "remediation", "recommendation", "notes"],
}


def _match_column(header: str) -> Optional[str]:
    """Given a raw header string, return the canonical field name or None."""
    h = header.strip().lower()
    for field, aliases in _COL_MAP.items():
        if h in aliases:
            return field
    return None


def parse_excel(file_bytes: bytes, filename: str = "") -> List[Dict[str, Any]]:
    """
    Parse an Excel workbook (.xlsx / .xls).

    Strategy:
      1. Read first sheet (or all sheets).
      2. Use the first row as headers; map them to canonical fields.
      3. Each subsequent row becomes one finding.
      4. Rows missing a title are skipped.
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    out: List[Dict[str, Any]] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue  # need at least header + 1 data row

        # Map header row
        raw_headers = [str(c or "").strip() for c in rows[0]]
        col_mapping: Dict[int, str] = {}
        for idx, h in enumerate(raw_headers):
            field = _match_column(h)
            if field:
                col_mapping[idx] = field

        # Must have at least a "title" column to proceed
        mapped_fields = set(col_mapping.values())
        if "title" not in mapped_fields:
            # Fallback: treat column 0 as title, column 1 as severity
            if len(raw_headers) >= 2:
                col_mapping = {0: "title", 1: "severity"}
                if len(raw_headers) >= 3:
                    col_mapping[2] = "cvss"
                if len(raw_headers) >= 4:
                    col_mapping[3] = "asset"
            else:
                continue

        for row in rows[1:]:
            record: Dict[str, Any] = {}
            for idx, field in col_mapping.items():
                val = row[idx] if idx < len(row) else None
                record[field] = val

            title = record.get("title")
            if not title or str(title).strip() == "":
                continue  # skip empty rows

            evidence_text = record.get("evidence", "")
            evidence_dict = {"description": str(evidence_text)} if evidence_text else {}

            raw_row = {raw_headers[i]: (row[i] if i < len(row) else None)
                       for i in range(len(raw_headers))}

            out.append(_make_finding(
                tool="excel",
                title=record.get("title"),
                severity=record.get("severity"),
                cvss=record.get("cvss"),
                asset=record.get("asset", ""),
                endpoint=record.get("endpoint"),
                cwe=str(record["cwe"]) if record.get("cwe") else None,
                owasp=str(record["owasp"]) if record.get("owasp") else None,
                evidence=evidence_dict,
                raw=raw_row,
            ))

    wb.close()
    return out


# ---------------------------------------------------------------------------
# 3. PDF parser
# ---------------------------------------------------------------------------

# Patterns for heuristic extraction of findings from PDF text
_SEV_RE = re.compile(
    r"\b(critical|high|medium|moderate|low|informational|info)\b",
    re.IGNORECASE,
)
_CVSS_RE = re.compile(r"\bCVSS[:\s]*(\d+\.?\d*)\b", re.IGNORECASE)
_CWE_RE = re.compile(r"\b(CWE-\d+)\b", re.IGNORECASE)
_OWASP_RE = re.compile(r"\b(A\d{1,2}:\d{4})\b")

# Common section headings that indicate a new finding
_FINDING_HEADER_RE = re.compile(
    r"^(?:\d+[\.\)]\s*|Finding\s*#?\d*[:\s]*|Vulnerability\s*#?\d*[:\s]*|Issue\s*#?\d*[:\s]*)"
    r"(.+)",
    re.IGNORECASE | re.MULTILINE,
)


def parse_pdf(file_bytes: bytes, filename: str = "") -> List[Dict[str, Any]]:
    """
    Parse a PDF file by extracting text and applying heuristics to identify
    vulnerability findings.

    Strategy:
      1. Extract full text with pdfplumber.
      2. Split into sections using heading patterns.
      3. For each section, extract severity, CVSS, CWE, OWASP.
      4. Build a finding per section.

    This is a best-effort parser. Real pentest PDFs vary enormously; the user
    can always export to Excel/JSON for higher fidelity.
    """
    import pdfplumber

    pdf = pdfplumber.open(io.BytesIO(file_bytes))
    full_text = ""
    for page in pdf.pages:
        page_text = page.extract_text()
        if page_text:
            full_text += page_text + "\n"
    pdf.close()

    if not full_text.strip():
        return []

    # --- Try structured extraction first (section-based) ---
    findings = _extract_findings_by_sections(full_text)

    # --- Fallback: table-like extraction ---
    if not findings:
        findings = _extract_findings_by_table_rows(full_text)

    # --- Last resort: whole-document single finding ---
    if not findings:
        sev_match = _SEV_RE.search(full_text)
        cvss_match = _CVSS_RE.search(full_text)
        cwe_match = _CWE_RE.search(full_text)
        owasp_match = _OWASP_RE.search(full_text)

        # Only create a finding if we found at least a severity or CVSS
        if sev_match or cvss_match:
            title = _guess_title_from_text(full_text) or f"Finding from {filename or 'PDF'}"
            findings.append(_make_finding(
                tool="pdf",
                title=title,
                severity=sev_match.group(1) if sev_match else "info",
                cvss=cvss_match.group(1) if cvss_match else None,
                cwe=cwe_match.group(1) if cwe_match else None,
                owasp=owasp_match.group(1) if owasp_match else None,
                evidence={"source_text": full_text[:2000]},
                raw={"filename": filename, "text_preview": full_text[:500]},
            ))

    return findings


def _extract_findings_by_sections(text: str) -> List[Dict[str, Any]]:
    """Split text by numbered/titled finding sections."""
    matches = list(_FINDING_HEADER_RE.finditer(text))
    if len(matches) < 1:
        return []

    findings: List[Dict[str, Any]] = []
    for i, m in enumerate(matches):
        title = m.group(1).strip()
        # Section body: from this match to the next, or end of text
        start = m.end()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(text)
        body = text[start:end].strip()

        sev_match = _SEV_RE.search(body)
        cvss_match = _CVSS_RE.search(body)
        cwe_match = _CWE_RE.search(body)
        owasp_match = _OWASP_RE.search(body)

        findings.append(_make_finding(
            tool="pdf",
            title=title,
            severity=sev_match.group(1) if sev_match else "info",
            cvss=cvss_match.group(1) if cvss_match else None,
            cwe=cwe_match.group(1) if cwe_match else None,
            owasp=owasp_match.group(1) if owasp_match else None,
            evidence={"description": body[:2000]},
            raw={"title": title, "body": body[:1000]},
        ))

    return findings


def _extract_findings_by_table_rows(text: str) -> List[Dict[str, Any]]:
    """
    Try to find table-like rows:
      severity | title-like text | CVSS
    Common in executive summaries.
    """
    # Pattern: a severity word at the start of a line, followed by tab/pipe/spaces
    row_re = re.compile(
        r"^(critical|high|medium|moderate|low|info(?:rmational)?)"
        r"[\s\t|]+(.+?)(?:[\s\t|]+(\d+\.?\d*))?$",
        re.IGNORECASE | re.MULTILINE,
    )
    findings: List[Dict[str, Any]] = []
    for m in row_re.finditer(text):
        sev = m.group(1)
        title = m.group(2).strip()
        cvss = m.group(3)
        if len(title) < 5 or len(title) > 300:
            continue

        findings.append(_make_finding(
            tool="pdf",
            title=title,
            severity=sev,
            cvss=cvss,
        ))

    return findings


def _guess_title_from_text(text: str) -> Optional[str]:
    """Try to extract a reasonable title from the first lines of text."""
    for line in text.split("\n")[:20]:
        line = line.strip()
        if 10 < len(line) < 120 and not line.startswith(("Page", "Date", "http")):
            return line
    return None


# ---------------------------------------------------------------------------
# 4. Router: pick the right parser based on file extension / content-type
# ---------------------------------------------------------------------------

def detect_and_parse(
    file_bytes: bytes,
    filename: str,
    content_type: str = "",
) -> List[Dict[str, Any]]:
    """
    Detect file format and delegate to the appropriate parser.
    Returns a list of normalized finding dicts.
    """
    fname = (filename or "").lower()
    ctype = (content_type or "").lower()

    # --- Excel ---
    if fname.endswith((".xlsx", ".xls")) or "spreadsheet" in ctype:
        return parse_excel(file_bytes, filename)

    # --- PDF ---
    if fname.endswith(".pdf") or ctype == "application/pdf":
        return parse_pdf(file_bytes, filename)

    # --- JSON (default) ---
    try:
        payload = json.loads(file_bytes.decode("utf-8"))
    except (json.JSONDecodeError, UnicodeDecodeError) as exc:
        raise ValueError(
            f"Unsupported file format: '{fname}'. "
            "Upload JSON, Excel (.xlsx), or PDF files."
        ) from exc

    return parse_generic_vuln_json(payload)
